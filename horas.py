# horas.py
# Requisitos:
#   pip install openpyxl
#   pip install "xlrd==1.2.0"   # solo si lees .xls de ORIGEN

import os, re, json, tempfile
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

CLASIF_JSON = "clasificacion_proyectos.json"   # guarda clasificaciones y exclusiones

# === Estilos ===
YELLOW = PatternFill("solid", fgColor="FFF4C2")  # Recurso
GREEN  = PatternFill("solid", fgColor="E6F4EA")  # Reparación
BLUE   = PatternFill("solid", fgColor="E3F2FD")  # Construcción
GRAY   = PatternFill("solid", fgColor="ECEFF1")  # Cabecera
BORDER = Border(left=Side(style="thin", color="D0D0D0"),
                right=Side(style="thin", color="D0D0D0"),
                top=Side(style="thin", color="D0D0D0"),
                bottom=Side(style="thin", color="D0D0D0"))
HDR_FONT = Font(bold=True)
TITLE_FONT = ("Segoe UI", 16, "bold")
SUBTITLE_FONT = ("Segoe UI", 9)

# === Compatibilidad .xls ===
try:
    import xlrd  # ==1.2.0
    XLRD_OK = True
except Exception:
    XLRD_OK = False

def xls_a_xlsx_si_hace_falta(path_in: str) -> str:
    if path_in.lower().endswith(".xlsx"):
        return path_in
    if not XLRD_OK:
        raise RuntimeError("El origen es .xls y no está xlrd==1.2.0. Instálalo o convierte a .xlsx.")
    if not os.path.exists(path_in):
        raise RuntimeError(f"Origen no encontrado: {path_in}")
    wb_xls = xlrd.open_workbook(path_in, formatting_info=False)
    tmpdir = tempfile.mkdtemp(prefix="xlsconv_")
    out = os.path.join(tmpdir, os.path.splitext(os.path.basename(path_in))[0] + "_conv.xlsx")
    wb_xlsx = Workbook()
    if wb_xlsx.active:
        wb_xlsx.remove(wb_xlsx.active)
    for si in range(wb_xls.nsheets):
        s_in = wb_xls.sheet_by_index(si)
        s_out = wb_xlsx.create_sheet(title=s_in.name)
        for r in range(s_in.nrows):
            for c in range(s_in.ncols):
                val = s_in.cell_value(r, c)
                if s_in.cell_type(r, c) == xlrd.XL_CELL_DATE:
                    try:
                        val = xlrd.xldate_as_datetime(val, wb_xls.datemode)
                    except Exception:
                        pass
                s_out.cell(row=r + 1, column=c + 1).value = val
    wb_xlsx.save(out)
    return out

# === Reglas de parseo ===
RE_PROY = re.compile(r'^\s*(\d{5})[ -](.+)$')  # proyecto = 5 dígitos
RE_MES_ANO = re.compile(r'(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+de\s+\d{4}', re.I)

def _norm(txt) -> str:
    s = "" if txt is None else str(txt)
    return re.sub(r'\s+', ' ', s).strip()

def _es_hdr_proyectos(txt: str) -> bool:
    return _norm(txt).lower().startswith("proyectos")

def normaliza_proyecto(texto) -> str:
    s = "" if texto is None else str(texto)
    m = RE_PROY.match(s)
    if not m: return ""
    return f"{m.group(1)}-{_norm(m.group(2))}"

def extraer_mapa_recurso_proyectos(ws):
    """Recurso = fila NO vacía inmediatamente anterior a 'Proyectos'."""
    mapa = {}
    recurso_actual = None
    for r in range(1, ws.max_row + 1):
        v = _norm(ws.cell(row=r, column=1).value)
        if not v:
            continue
        if _es_hdr_proyectos(v):
            k = r - 1
            prev = ""
            while k >= 1 and not prev:
                prev = _norm(ws.cell(row=k, column=1).value)
                k -= 1
            if prev:
                recurso_actual = prev
                mapa.setdefault(recurso_actual, [])
            continue
        if RE_PROY.match(v):
            p = normaliza_proyecto(v)
            if not recurso_actual:
                mapa.setdefault("_SIN_RECURSO_", [])
                if p not in mapa["_SIN_RECURSO_"]:
                    mapa["_SIN_RECURSO_"].append(p)
            else:
                if p not in mapa[recurso_actual]:
                    mapa[recurso_actual].append(p)
    return mapa

def extraer_proyectos(ws):
    proys = set()
    for lst in extraer_mapa_recurso_proyectos(ws).values():
        for p in lst: proys.add(p)
    return sorted(proys)

# === Encabezado único ===
def encontrar_bloque_encabezado(ws):
    """Devuelve (fila_inicio, fila_fin) del bloque de encabezado superior que solo se copiará una vez."""
    fila_recurso = None
    for r in range(1, min(ws.max_row, 120)+1):
        v1 = _norm(ws.cell(row=r, column=1).value)
        if v1.lower() == "recurso":
            fila_recurso = r
            break
    if fila_recurso is None:
        return (0, 0)
    start = max(1, fila_recurso-2)
    for k in range(fila_recurso-3, max(0, fila_recurso-12), -1):
        txt = " ".join(_norm(ws.cell(row=k, column=c).value) for c in range(1, 6))
        if RE_MES_ANO.search(txt):
            start = k
            break
    return (start, fila_recurso)

# === Localizar TOTAL y conversiones ===
def localizar_columna_total(ws, scan_rows=25):
    for r in range(1, min(ws.max_row, scan_rows) + 1):
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(row=r, column=c).value or "").strip().upper() == "TOTAL":
                return c
    return None

def hora_a_minutos(valor):
    if valor is None or str(valor).strip() == "": return 0
    # acepta 08:00, 8:00, 8, 8.5, "8,5"
    try:
        return int(float(str(valor).replace(",", ".")) * 60)
    except Exception:
        pass
    s = _norm(valor)
    m = re.match(r"^\s*(\d+)\s*[:\.]\s*(\d{1,2})\s*$", s)
    if m:
        return int(m.group(1)) * 60 + int(m.group(2))
    return 0

def minutos_a_hhmm(mins:int) -> str:
    h = mins // 60
    m = mins % 60
    return f"{h:02d}:{m:02d}"

def hora_a_decimal(valor):
    return round(hora_a_minutos(valor) / 60.0, 2)

# === Estilado y copia ===
def _style_row(ws, r, fill=None, bold=False):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=r, column=c)
        if fill: cell.fill = fill
        cell.border = BORDER
        if bold: cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center")

def _copy_row(ws_src, ws_dst, r_src, r_dst, max_cols=None):
    if max_cols is None: max_cols = ws_src.max_column
    for c in range(1, min(ws_src.max_column, max_cols) + 1):
        ws_dst.cell(row=r_dst, column=c).value = ws_src.cell(row=r_src, column=c).value

# === Construcción del libro de salida ===
def construir_wb_salida(ws_src, clasif_dict, excl_recursos:set, excl_proyectos:set):
    wb_out = Workbook(); ws_out = wb_out.active; ws_out.title = "IPI"

    # Anchos de columna aproximados
    for col_letter, dim in ws_src.column_dimensions.items():
        ws_out.column_dimensions[col_letter].width = dim.width or 14

    # Copiar solo una vez el encabezado superior
    hdr_ini, hdr_fin = encontrar_bloque_encabezado(ws_src)
    dst = 1
    if hdr_fin > 0:
        for r in range(hdr_ini, hdr_fin + 1):
            _copy_row(ws_src, ws_out, r, dst)
            _style_row(ws_out, dst, fill=GRAY, bold=True)
            dst += 1

    recurso_actual = None
    excluir_bloque = False

    # Recorrer todo el sheet de origen
    r = 1
    while r <= ws_src.max_row:
        v = _norm(ws_src.cell(row=r, column=1).value)

        if _es_hdr_proyectos(v):
            # Recurso justo encima
            k = r - 1
            prev = ""
            while k >= 1 and not prev:
                prev = _norm(ws_src.cell(row=k, column=1).value)
                k -= 1
            recurso_actual = prev or None
            excluir_bloque = recurso_actual in excl_recursos if recurso_actual else False

            if recurso_actual and not excluir_bloque:
                ws_out.cell(row=dst, column=1).value = recurso_actual
                _style_row(ws_out, dst, fill=YELLOW, bold=True)
                dst += 1

            # Copiar la fila "Proyectos:"
            if not excluir_bloque:
                _copy_row(ws_src, ws_out, r, dst)
                _style_row(ws_out, dst, fill=GRAY, bold=True)
                dst += 1

            r += 1
            continue

        if excluir_bloque:
            r += 1
            continue

        # Proyecto u otras filas
        if RE_PROY.match(v):
            p = normaliza_proyecto(v)
            if p in excl_proyectos:
                r += 1; continue

            _copy_row(ws_src, ws_out, r, dst)
            tipo = clasif_dict.get(p, None)
            if tipo == "REPARACION":
                _style_row(ws_out, dst, fill=GREEN)
            elif tipo == "CONSTRUCCION":
                _style_row(ws_out, dst, fill=BLUE)
            else:
                _style_row(ws_out, dst)
            dst += 1
        else:
            # Copia “1-HORA NORMAL”, “2-HORA EXTRA”, etc. y cualquier otra fila del bloque
            _copy_row(ws_src, ws_out, r, dst)
            _style_row(ws_out, dst)
            dst += 1

        r += 1

    # Localizar TOTAL en salida y recalcular totales por fila
    col_total = localizar_columna_total(ws_out, scan_rows=hdr_fin+10 if hdr_fin else 25)
    if not col_total:
        # si no existe, crea cabecera TOTAL a continuación de la última columna usada en la cabecera
        col_total = ws_out.max_column + 1
        ws_out.cell(row=hdr_fin if hdr_fin else 1, column=col_total, value="TOTAL")
        _style_row(ws_out, hdr_fin if hdr_fin else 1, fill=GRAY, bold=True)

    # Crear columna TOTAL DEC
    ws_out.cell(row=(hdr_fin if hdr_fin else 1), column=col_total + 1, value="TOTAL DEC")
    _style_row(ws_out, (hdr_fin if hdr_fin else 1), fill=GRAY, bold=True)

    # Acumuladores por tipo de imputación usando TOTAL DEC de filas de proyecto
    acum_rep = 0.0
    acum_con = 0.0

    # Recalcular cada fila de datos: suma todas las celdas de horas entre la columna 3 y TOTAL-1
    first_data_row = (hdr_fin + 1) if hdr_fin else 2
    for rr in range(first_data_row, ws_out.max_row + 1):
        c1 = _norm(ws_out.cell(row=rr, column=1).value)
        # detectar filas de “Proyectos:” y encabezados recurso
        if _es_hdr_proyectos(c1) or c1.endswith(":") or c1 == "" or c1 is None:
            continue

        # Sumar horas de la fila
        mins = 0
        for cc in range(3, col_total):   # col 1 = proyecto/título, col 2 = tipo
            mins += hora_a_minutos(ws_out.cell(row=rr, column=cc).value)

        # Escribir TOTAL hh:mm y TOTAL DEC
        ws_out.cell(row=rr, column=col_total, value=minutos_a_hhmm(mins) if mins else None)
        dec = round(mins / 60.0, 2)
        ws_out.cell(row=rr, column=col_total + 1, value=dec if mins else None)

        # Acumular por tipo solo si es fila de proyecto
        if RE_PROY.match(c1):
            p = normaliza_proyecto(c1)
            tipo = clasif_dict.get(p)
            if tipo == "REPARACION":
                acum_rep += dec
            elif tipo == "CONSTRUCCION":
                acum_con += dec

    # Sección de totales por tipo
    fila_sum = ws_out.max_row + 2
    ws_out.cell(row=fila_sum, column=1, value="Totales por tipo de imputación")
    _style_row(ws_out, fila_sum, fill=GRAY, bold=True)
    ws_out.merge_cells(start_row=fila_sum, start_column=1, end_row=fila_sum, end_column=max(2, col_total + 1))

    ws_out.cell(row=fila_sum + 1, column=1, value="REPARACIÓN")
    ws_out.cell(row=fila_sum + 1, column=2, value=round(acum_rep, 2))
    _style_row(ws_out, fila_sum + 1, fill=GREEN, bold=True)

    ws_out.cell(row=fila_sum + 2, column=1, value="CONSTRUCCIÓN")
    ws_out.cell(row=fila_sum + 2, column=2, value=round(acum_con, 2))
    _style_row(ws_out, fila_sum + 2, fill=BLUE, bold=True)

    return wb_out

# === Clasificaciones y exclusiones persistentes ===
def cargar_estado():
    if not os.path.exists(CLASIF_JSON):
        return {"clasif": {}, "excl": {"recursos": [], "proyectos": []}}
    try:
        with open(CLASIF_JSON, "r", encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, dict) and "clasif" in data and "excl" in data:
                return data
            # compat antiguo: solo dict de clasif
            return {"clasif": data if isinstance(data, dict) else {}, "excl": {"recursos": [], "proyectos": []}}
    except Exception:
        return {"clasif": {}, "excl": {"recursos": [], "proyectos": []}}

def guardar_estado(clasif_dict, excl_recursos, excl_proyectos):
    data = {"clasif": clasif_dict, "excl": {"recursos": sorted(excl_recursos), "proyectos": sorted(excl_proyectos)}}
    with open(CLASIF_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# === Diálogos ===
class ClasificarDialog(tk.Toplevel):
    def __init__(self, master, proyectos_nuevos):
        super().__init__(master)
        self.title("Clasificar proyectos nuevos")
        self.geometry("980x600")
        self.resizable(True, True)
        self.result = None

        ttk.Label(self, text=f"Clasifica los proyectos. Se guardará en {os.path.basename(CLASIF_JSON)}.").pack(
            fill="x", padx=16, pady=(14,6)
        )

        container = ttk.Frame(self); container.pack(fill="both", expand=True, padx=16, pady=10)
        left = ttk.Frame(container); left.pack(side="left", fill="both", expand=True)
        controls = ttk.Frame(container); controls.pack(side="left", fill="y", padx=8)
        right = ttk.Frame(container); right.pack(side="left", fill="both", expand=True)

        self.var_search = tk.StringVar()
        ttk.Label(left, text="Nuevos").pack(anchor="w")
        ttk.Entry(left, textvariable=self.var_search).pack(fill="x", pady=(0,6))
        self.var_search.trace_add("write", lambda *_: self._apply_filter())

        self.listbox = tk.Listbox(left, selectmode=tk.EXTENDED)
        self.listbox.pack(fill="both", expand=True)
        self._all_items = sorted(proyectos_nuevos)
        for p in self._all_items: self.listbox.insert(tk.END, p)

        ttk.Button(controls, text="→ CONSTRUCCIÓN", command=self._mark_construccion).pack(fill="x", pady=4)
        ttk.Button(controls, text="→ REPARACIÓN", command=self._mark_reparacion).pack(fill="x", pady=4)
        ttk.Separator(controls, orient="horizontal").pack(fill="x", pady=12)
        ttk.Button(controls, text="Limpiar selección", command=lambda: self.listbox.selection_clear(0, tk.END)).pack(fill="x")

        ttk.Label(right, text="CONSTRUCCIÓN").pack()
        self.list_constr = tk.Listbox(right); self.list_constr.pack(fill="both", expand=True, padx=4, pady=4)
        ttk.Label(right, text="REPARACIÓN").pack()
        self.list_repara = tk.Listbox(right); self.list_repara.pack(fill="both", expand=True, padx=4, pady=4)

        bottom = ttk.Frame(self); bottom.pack(fill="x", padx=16, pady=12)
        ttk.Button(bottom, text="Guardar", command=self._save).pack(side="right", padx=6)
        ttk.Button(bottom, text="Cancelar", command=self.destroy).pack(side="right")

    def _apply_filter(self):
        q = self.var_search.get().strip().lower()
        self.listbox.delete(0, tk.END)
        for p in self._all_items:
            if q in p.lower():
                self.listbox.insert(tk.END, p)

    def _selected_items(self):
        return [self.listbox.get(i) for i in self.listbox.curselection()]

    def _remove_from(self, lb: tk.Listbox, value: str):
        items = lb.get(0, tk.END)
        for i, v in enumerate(items):
            if v == value:
                lb.delete(i); break

    def _mark_construccion(self):
        for p in self._selected_items():
            if p not in self.list_constr.get(0, tk.END):
                self.list_constr.insert(tk.END, p)
            self._remove_from(self.list_repara, p)

    def _mark_reparacion(self):
        for p in self._selected_items():
            if p not in self.list_repara.get(0, tk.END):
                self.list_repara.insert(tk.END, p)
            self._remove_from(self.list_constr, p)

    def _save(self):
        res = {}
        for v in self.list_constr.get(0, tk.END): res[v] = "CONSTRUCCION"
        for v in self.list_repara.get(0, tk.END): res[v] = "REPARACION"
        if not res:
            messagebox.showwarning("Atención", "No has clasificado ningún proyecto.")
            return
        self.result = res
        self.destroy()

class EditarClasifDialog(tk.Toplevel):
    def __init__(self, master, clasif_dict):
        super().__init__(master)
        self.title("Editar clasificaciones")
        self.geometry("1000x640")
        self.resizable(True, True)
        self.res = None

        rootf = ttk.Frame(self, padding=10)
        rootf.pack(fill="both", expand=True)
        rootf.columnconfigure(0, weight=1)
        rootf.columnconfigure(1, weight=0)
        rootf.columnconfigure(2, weight=1)
        rootf.rowconfigure(3, weight=1)

        ttk.Label(rootf, text="CONSTRUCCIÓN").grid(row=0, column=0, sticky="w")
        ttk.Label(rootf, text="REPARACIÓN").grid(row=0, column=2, sticky="w")

        self.var_c = tk.StringVar(); self.var_r = tk.StringVar()
        ttk.Entry(rootf, textvariable=self.var_c).grid(row=1, column=0, sticky="ew", pady=(0,6))
        ttk.Entry(rootf, textvariable=self.var_r).grid(row=1, column=2, sticky="ew", pady=(0,6))

        self.lb_c = tk.Listbox(rootf, selectmode=tk.EXTENDED)
        self.lb_r = tk.Listbox(rootf, selectmode=tk.EXTENDED)
        self.lb_c.grid(row=3, column=0, sticky="nsew", padx=(0,6))
        self.lb_r.grid(row=3, column=2, sticky="nsew", padx=(6,0))

        mid = ttk.Frame(rootf); mid.grid(row=3, column=1, sticky="ns")
        ttk.Button(mid, text="→ Reparación", command=self._to_r).pack(pady=6)
        ttk.Button(mid, text="→ Construcción", command=self._to_c).pack(pady=6)

        self._all_c = sorted([p for p, t in clasif_dict.items() if t == "CONSTRUCCION"])
        self._all_r = sorted([p for p, t in clasif_dict.items() if t == "REPARACION"])
        for v in self._all_c: self.lb_c.insert(tk.END, v)
        for v in self._all_r: self.lb_r.insert(tk.END, v)

        self.var_c.trace_add("write", lambda *_: self._filter(self.lb_c, self._all_c, self.var_c.get()))
        self.var_r.trace_add("write", lambda *_: self._filter(self.lb_r, self._all_r, self.var_r.get()))

        bottom = ttk.Frame(self); bottom.pack(side="bottom", fill="x", padx=10, pady=10)
        ttk.Button(bottom, text="Guardar", command=self._save).pack(side="right", padx=6)
        ttk.Button(bottom, text="Cancelar", command=self.destroy).pack(side="right")

    def _filter(self, lb, items, q):
        q = (q or "").lower().strip()
        lb.delete(0, tk.END)
        for it in items:
            if q in it.lower():
                lb.insert(tk.END, it)

    def _move(self, src, dst, src_all, dst_all):
        idxs = list(src.curselection()); idxs.reverse()
        moved = []
        for i in idxs:
            v = src.get(i)
            moved.append(v)
            src.delete(i)
        for v in moved:
            if v not in dst.get(0, tk.END):
                dst.insert(tk.END, v)
            if v in src_all: src_all.remove(v)
            if v not in dst_all: dst_all.append(v)
        self._filter(src, src_all, (self.var_c.get() if src is self.lb_c else self.var_r.get()))
        self._filter(dst, dst_all, (self.var_r.get() if dst is self.lb_r else self.var_c.get()))

    def _to_r(self): self._move(self.lb_c, self.lb_r, self._all_c, self._all_r)
    def _to_c(self): self._move(self.lb_r, self.lb_c, self._all_r, self._all_c)

    def _save(self):
        out = {}
        for v in self.lb_c.get(0, tk.END): out[v] = "CONSTRUCCION"
        for v in self.lb_r.get(0, tk.END): out[v] = "REPARACION"
        self.res = out
        self.destroy()

class ExcluirDialog(tk.Toplevel):
    def __init__(self, master, recursos:list, proyectos:list, excl_r:set, excl_p:set):
        super().__init__(master)
        self.title("Excluir recursos y/o proyectos")
        self.geometry("1100x620")
        self.resizable(True, True)
        self.excl_recursos = set(excl_r)
        self.excl_proyectos = set(excl_p)

        ttk.Label(self, text="Selecciona elementos a EXCLUIR del cálculo. Deja vacío si no quieres excluir.").pack(
            fill="x", padx=16, pady=(14,6)
        )

        wrap = ttk.Frame(self); wrap.pack(fill="both", expand=True, padx=16, pady=10)

        fr = ttk.Labelframe(wrap, text="Recursos"); fr.pack(side="left", fill="both", expand=True, padx=(0,8))
        self.var_r = tk.StringVar(); ttk.Entry(fr, textvariable=self.var_r).pack(fill="x", padx=8, pady=6)
        self.lb_r = tk.Listbox(fr, selectmode=tk.EXTENDED); self.lb_r.pack(fill="both", expand=True, padx=8, pady=(0,8))
        self._all_rec = sorted(recursos)
        for x in self._all_rec: self.lb_r.insert(tk.END, x)
        self.var_r.trace_add("write", lambda *_: self._filter_list(self.lb_r, self._all_rec, self.var_r.get()))

        fp = ttk.Labelframe(wrap, text="Proyectos"); fp.pack(side="left", fill="both", expand=True, padx=(8,0))
        self.var_p = tk.StringVar(); ttk.Entry(fp, textvariable=self.var_p).pack(fill="x", padx=8, pady=6)
        self.lb_p = tk.Listbox(fp, selectmode=tk.EXTENDED); self.lb_p.pack(fill="both", expand=True, padx=8, pady=(0,8))
        self._all_proj = sorted(proyectos)
        for x in self._all_proj: self.lb_p.insert(tk.END, x)
        self.var_p.trace_add("write", lambda *_: self._filter_list(self.lb_p, self._all_proj, self.var_p.get()))

        bottom = ttk.Frame(self); bottom.pack(fill="x", padx=16, pady=12)
        ttk.Button(bottom, text="Aplicar exclusiones", command=self._apply).pack(side="right", padx=6)
        ttk.Button(bottom, text="No excluir", command=self._skip).pack(side="right")

        # preselección con lo que había en memoria
        for i, v in enumerate(self._all_rec):
            if v in self.excl_recursos: self.lb_r.selection_set(i)
        for i, v in enumerate(self._all_proj):
            if v in self.excl_proyectos: self.lb_p.selection_set(i)

    def _filter_list(self, lb: tk.Listbox, items: list, q: str):
        q = (q or "").lower().strip()
        lb.delete(0, tk.END)
        for it in items:
            if q in it.lower():
                lb.insert(tk.END, it)

    def _apply(self):
        self.excl_recursos = set(self.lb_r.get(i) for i in self.lb_r.curselection())
        self.excl_proyectos = set(self.lb_p.get(i) for i in self.lb_p.curselection())
        self.destroy()

    def _skip(self):
        self.destroy()

# === Transformación ===
def transformar(origen_path: str, clasif_dict: dict,
                excl_recursos:set, excl_proyectos:set) -> str:
    in_resolved = xls_a_xlsx_si_hace_falta(origen_path)
    wb_in = load_workbook(in_resolved, data_only=True)
    ws_in = wb_in.active

    # Clasificar nuevos
    proyectos_all = extraer_proyectos(ws_in)
    nuevos = [p for p in proyectos_all if p not in clasif_dict]
    if nuevos:
        dlg = ClasificarDialog(root, nuevos)
        dlg.grab_set(); root.wait_window(dlg)
        if not dlg.result:
            raise RuntimeError("Clasificación cancelada. No se puede continuar.")
        clasif_dict.update(dlg.result)
        guardar_estado(clasif_dict, excl_recursos, excl_proyectos)

    wb_out = construir_wb_salida(ws_in, clasif_dict, excl_recursos, excl_proyectos)

    # Salida en la misma carpeta del origen
    base_dir = os.path.dirname(origen_path)
    base_name = os.path.splitext(os.path.basename(origen_path))[0]
    out_path = os.path.join(base_dir, f"{base_name}_IPI.xlsx")
    wb_out.save(out_path)
    return out_path

# === App (UI) ===
class App:
    def __init__(self, root):
        self.root = root
        root.title("Astillero · Transformador IPI")
        root.geometry("1000x680"); root.minsize(900, 600)

        style = ttk.Style()
        try: style.theme_use("clam")
        except Exception: pass
        style.configure("TButton", padding=8)
        style.configure("Accent.TButton", padding=10, font=("Segoe UI", 10, "bold"))
        style.map("Accent.TButton", background=[("active", "#0b6efd")])
        style.configure("Card.TLabelframe", background="#ffffff")
        style.configure("Card.TLabelframe.Label", font=("Segoe UI", 10, "bold"))
        style.configure("Small.TLabel", foreground="#6b7280")
        style.configure("Banner.TFrame", background="#0f172a")

        banner = ttk.Frame(root, style="Banner.TFrame")
        banner.pack(fill="x")
        tk.Label(banner, text="Transformador Excel → IPI",
                 font=TITLE_FONT, fg="#ffffff", bg="#0f172a").pack(anchor="w", padx=18, pady=(12,0))
        tk.Label(banner, text="Colores y totales por tipo. Salida junto al origen.",
                 font=SUBTITLE_FONT, fg="#cbd5e1", bg="#0f172a").pack(anchor="w", padx=18, pady=(0,12))

        self.in_path = None
        self.out_preview = tk.StringVar(value="—")
        estado = cargar_estado()
        self.clasif = estado["clasif"]
        self.excl_recursos = set(estado["excl"].get("recursos", []))
        self.excl_proyectos = set(estado["excl"].get("proyectos", []))

        main = ttk.Frame(root, padding=(16, 12)); main.pack(fill="both", expand=True)

        fsel = ttk.Labelframe(main, text="Archivos", style="Card.TLabelframe")
        fsel.pack(fill="x", pady=8)

        row = ttk.Frame(fsel); row.pack(fill="x", pady=8, padx=10)
        ttk.Label(row, text="Origen (.xlsx/.xls):", width=22).pack(side="left")
        self.lbl_in = ttk.Label(row, text="Sin archivo", style="Small.TLabel")
        self.lbl_in.pack(side="left", fill="x", expand=True)
        ttk.Button(row, text="Seleccionar…", command=self.sel_in).pack(side="left", padx=8)

        row2 = ttk.Frame(fsel); row2.pack(fill="x", pady=8, padx=10)
        ttk.Label(row2, text="Salida prevista:", width=22).pack(side="left")
        ttk.Label(row2, textvariable=self.out_preview).pack(side="left", fill="x", expand=True)

        actions = ttk.Frame(main); actions.pack(fill="x", pady=10)
        ttk.Button(actions, text="Editar clasificaciones…", command=self.editar_clasificaciones).pack(side="left")
        self.btn_excl = ttk.Button(actions, text="Exclusiones…", command=self.pedir_exclusiones, state="disabled")
        self.btn_excl.pack(side="left", padx=6)
        self.btn_run = ttk.Button(actions, text="Transformar", style="Accent.TButton",
                                  command=self.run, state="disabled")
        self.btn_run.pack(side="right")

        cons = ttk.Labelframe(main, text="Registro", style="Card.TLabelframe")
        cons.pack(fill="both", expand=True, pady=(6, 12))
        self.txt = tk.Text(cons, height=18, bg="#f8fafc", bd=0, highlightthickness=0)
        self.txt.pack(fill="both", expand=True, padx=10, pady=10)

        status = ttk.Frame(root); status.pack(fill="x")
        ttk.Label(status, text=f"Proyectos clasificados: {len(self.clasif)}", style="Small.TLabel").pack(side="right", padx=12, pady=6)

    def log(self, m):
        self.txt.insert(tk.END, m + "\n")
        self.txt.see(tk.END)
        self.root.update_idletasks()

    def sel_in(self):
        p = filedialog.askopenfilename(title="Selecciona Excel origen", filetypes=[("Excel", "*.xlsx *.xls")])
        if p:
            self.in_path = p
            self.lbl_in.config(text=os.path.basename(p))
            base_dir = os.path.dirname(p)
            base_name = os.path.splitext(os.path.basename(p))[0]
            self.out_preview.set(os.path.join(base_dir, f"{base_name}_IPI.xlsx"))
            self._check_ready()

    def _check_ready(self):
        ok = bool(self.in_path)
        state = "normal" if ok else "disabled"
        self.btn_run.config(state=state)
        self.btn_excl.config(state=state)

    def editar_clasificaciones(self):
        if not self.clasif:
            messagebox.showinfo("Info", "Aún no hay clasificaciones guardadas.")
            return
        dlg = EditarClasifDialog(self.root, self.clasif)
        dlg.grab_set(); self.root.wait_window(dlg)
        if dlg.res is not None:
            self.clasif = dlg.res
            guardar_estado(self.clasif, self.excl_recursos, self.excl_proyectos)
            messagebox.showinfo("Listo", "Clasificaciones actualizadas.")

    def pedir_exclusiones(self):
        try:
            in_resolved = xls_a_xlsx_si_hace_falta(self.in_path)
            wb_in = load_workbook(in_resolved, data_only=True)
            ws_in = wb_in.active
            mapa = extraer_mapa_recurso_proyectos(ws_in)
            recursos = [k for k in mapa.keys() if k != "_SIN_RECURSO_"]
            proyectos = extraer_proyectos(ws_in)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el archivo de origen para preparar exclusiones.\n{e}")
            return

        dlg = ExcluirDialog(self.root, recursos, proyectos, self.excl_recursos, self.excl_proyectos)
        dlg.grab_set(); self.root.wait_window(dlg)
        self.excl_recursos = dlg.excl_recursos
        self.excl_proyectos = dlg.excl_proyectos
        guardar_estado(self.clasif, self.excl_recursos, self.excl_proyectos)
        self.log(f"Exclusiones: {len(self.excl_recursos)} recursos, {len(self.excl_proyectos)} proyectos.")

    def run(self):
        try:
            self.btn_run.config(state="disabled")
            self.txt.delete("1.0", tk.END)
            out_path = transformar(self.in_path, self.clasif, self.excl_recursos, self.excl_proyectos)
            self.log(f"Guardado: {out_path}")
            messagebox.showinfo("Listo", f"Archivo creado:\n{out_path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))
        finally:
            self.btn_run.config(state="normal")

# === main ===
if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()
